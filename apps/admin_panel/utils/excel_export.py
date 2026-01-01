"""
Utility functions for exporting reports to Excel format.
"""
import os
from io import BytesIO
from typing import Dict, Any, List, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.conf import settings
from django.utils import timezone
from urllib.parse import quote


class ExcelExporter:
    """Utility class for exporting data to Excel format."""
    
    # رنگ‌های هدر
    HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
    
    # استایل سلول‌های عادی
    NORMAL_FONT = Font(size=10)
    NORMAL_ALIGNMENT = Alignment(horizontal="right", vertical="center")
    
    # استایل خلاصه آماری
    SUMMARY_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    SUMMARY_FONT = Font(bold=True, size=11)
    SUMMARY_LABEL_FONT = Font(bold=True, size=10)
    
    @staticmethod
    def translate_status(status: str) -> str:
        """Translate order status to Persian."""
        status_map = {
            'pending': 'در انتظار',
            'processing': 'در حال پردازش',
            'completed': 'تکمیل شده',
            'cancelled': 'لغو شده',
            'refunded': 'بازگشت شده',
        }
        return status_map.get(status, status)
    
    @staticmethod
    def translate_payment_status(status: str) -> str:
        """Translate payment status to Persian."""
        status_map = {
            'pending': 'در انتظار',
            'paid': 'پرداخت شده',
            'failed': 'ناموفق',
            'refunded': 'بازگشت شده',
            'cancelled': 'لغو شده',
        }
        return status_map.get(status, status)
    
    @staticmethod
    def translate_payment_method(method: str) -> str:
        """Translate payment method to Persian."""
        method_map = {
            'cash': 'نقدی',
            'card': 'کارت',
            'pos': 'دستگاه پوز',
            'online': 'آنلاین',
        }
        return method_map.get(method, method) if method else ''
    
    @staticmethod
    def create_workbook() -> Workbook:
        """Create a new Excel workbook."""
        return Workbook()
    
    @staticmethod
    def style_header_row(worksheet, row_num: int, num_cols: int):
        """Apply styling to header row."""
        for col in range(1, num_cols + 1):
            cell = worksheet.cell(row=row_num, column=col)
            cell.fill = ExcelExporter.HEADER_FILL
            cell.font = ExcelExporter.HEADER_FONT
            cell.alignment = ExcelExporter.HEADER_ALIGNMENT
    
    @staticmethod
    def style_data_row(worksheet, row_num: int, num_cols: int):
        """Apply styling to data row."""
        for col in range(1, num_cols + 1):
            cell = worksheet.cell(row=row_num, column=col)
            cell.font = ExcelExporter.NORMAL_FONT
            cell.alignment = ExcelExporter.NORMAL_ALIGNMENT
    
    @staticmethod
    def auto_adjust_column_width(worksheet):
        """Automatically adjust column widths."""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def save_excel_to_media(workbook: Workbook, filename: str, request=None) -> str:
        """
        Save Excel workbook to media directory and return full URL.
        
        Args:
            workbook: OpenPyXL Workbook object
            filename: Filename for the Excel file
            request: HTTP request object (optional, for building full URL)
            
        Returns:
            str: Full URL to access the file
        """
        # Create reports directory in media if it doesn't exist
        reports_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
        os.makedirs(reports_dir, exist_ok=True)
        
        # Full file path
        file_path = os.path.join(reports_dir, filename)
        
        # Save workbook to file
        workbook.save(file_path)
        
        # Generate full URL
        if request:
            # Use request to build absolute URL
            file_url = request.build_absolute_uri(f"{settings.MEDIA_URL}reports/{filename}")
        else:
            # Fallback to relative URL
            file_url = f"{settings.MEDIA_URL}reports/{filename}"
        
        return file_url
    
    @staticmethod
    def create_excel_response(workbook: Workbook, default_filename: str) -> HttpResponse:
        """
        Create HTTP response for Excel file download (legacy method).
        
        Args:
            workbook: OpenPyXL Workbook object
            default_filename: Default filename if not provided
            
        Returns:
            HttpResponse: HTTP response with Excel file
        """
        # Save workbook to BytesIO
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        
        # Encode filename for proper handling of Persian characters
        encoded_filename = quote(default_filename.encode('utf-8'))
        
        # Create response
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{encoded_filename}'
        response['Content-Length'] = len(response.content)
        
        output.close()
        return response
    
    @staticmethod
    def export_sales_report(report_data: Dict[str, Any], filename: str = None, request=None) -> str:
        """Export sales report to Excel with complete data and statistics."""
        wb = ExcelExporter.create_workbook()
        ws = wb.active
        ws.title = "گزارش فروش"
        
        # خلاصه آماری در ابتدا
        row_num = 1
        ws.merge_cells(f'A{row_num}:B{row_num}')
        summary_cell = ws.cell(row=row_num, column=1, value="خلاصه آماری گزارش فروش")
        summary_cell.fill = ExcelExporter.SUMMARY_FILL
        summary_cell.font = ExcelExporter.SUMMARY_FONT
        summary_cell.alignment = Alignment(horizontal="center", vertical="center")
        row_num += 1
        
        # آمار کلی (شامل آمار فروش و تراکنش)
        total_transactions = report_data.get('total_transactions', 0)
        successful = report_data.get('successful_transactions', 0)
        failed = report_data.get('failed_transactions', 0)
        success_rate = (successful / total_transactions * 100) if total_transactions > 0 else 0
        
        stats = [
            ("تاریخ شروع:", report_data.get('start_date', 'همه تاریخ‌ها')),
            ("تاریخ پایان:", report_data.get('end_date', 'همه تاریخ‌ها')),
            ("", ""),
            ("=== آمار فروش ===", ""),
            ("مجموع فروش (ریال):", f"{report_data.get('total_sales', 0):,}"),
            ("تعداد کل سفارشات:", report_data.get('total_orders', 0)),
            ("میانگین ارزش هر سفارش (ریال):", f"{report_data.get('average_order_value', 0):,.2f}"),
            ("", ""),
            ("=== آمار تراکنش‌ها ===", ""),
            ("تعداد کل تراکنش‌ها:", total_transactions),
            ("تراکنش‌های موفق:", successful),
            ("تراکنش‌های ناموفق:", failed),
            ("نرخ موفقیت (%):", f"{success_rate:.2f}"),
            ("مجموع مبلغ موفق (ریال):", f"{report_data.get('successful_amount', 0):,}"),
        ]
        
        for label, value in stats:
            ws.cell(row=row_num, column=1, value=label).font = ExcelExporter.SUMMARY_LABEL_FONT
            ws.cell(row=row_num, column=2, value=value).font = ExcelExporter.NORMAL_FONT
            row_num += 1
        
        row_num += 1  # فاصله
        
        # هدر جدول
        headers = [
            "شناسه", "شماره سفارش", "مبلغ کل (ریال)", "وضعیت", 
            "وضعیت پرداخت", "شناسه تراکنش", "Gateway", "روش پرداخت",
            "تاریخ ایجاد", "تاریخ بروزرسانی"
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=header)
            cell.fill = ExcelExporter.HEADER_FILL
            cell.font = ExcelExporter.HEADER_FONT
            cell.alignment = ExcelExporter.HEADER_ALIGNMENT
        row_num += 1
        
        # داده‌های کامل
        orders = report_data.get('orders', [])
        for order in orders:
            row = [
                order.get('id', ''),
                order.get('order_number', ''),
                order.get('total_amount', 0),
                order.get('status', ''),
                order.get('payment_status', ''),
                order.get('transaction_id', ''),
                order.get('gateway_name', ''),
                order.get('payment_method', ''),
                str(order.get('created_at', ''))[:19] if order.get('created_at') else '',
                str(order.get('updated_at', ''))[:19] if order.get('updated_at') else ''
            ]
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.font = ExcelExporter.NORMAL_FONT
                cell.alignment = ExcelExporter.NORMAL_ALIGNMENT
            row_num += 1
        
        ExcelExporter.auto_adjust_column_width(ws)
        
        # ذخیره فایل در media و برگرداندن لینک
        if not filename:
            filename = f"sales_report_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        file_url = ExcelExporter.save_excel_to_media(wb, filename, request)
        return file_url
    
    @staticmethod
    def export_product_report(report_data: Dict[str, Any], filename: str = None, request=None) -> str:
        """Export product report to Excel with complete data and statistics."""
        wb = ExcelExporter.create_workbook()
        ws = wb.active
        ws.title = "گزارش محصولات"
        
        # خلاصه آماری در ابتدا
        row_num = 1
        ws.merge_cells(f'A{row_num}:B{row_num}')
        summary_cell = ws.cell(row=row_num, column=1, value="خلاصه آماری گزارش محصولات")
        summary_cell.fill = ExcelExporter.SUMMARY_FILL
        summary_cell.font = ExcelExporter.SUMMARY_FONT
        summary_cell.alignment = Alignment(horizontal="center", vertical="center")
        row_num += 1
        
        total_products = report_data.get('total_products', 0)
        active_products = report_data.get('active_products', 0)
        products = report_data.get('products', [])
        total_revenue = sum(p.get('total_revenue') or 0 for p in products)
        total_sold = sum(p.get('total_sold') or 0 for p in products)
        
        # آمار کلی
        stats = [
            ("تعداد کل محصولات:", total_products),
            ("محصولات فعال:", active_products),
            ("محصولات غیرفعال:", total_products - active_products),
            ("", ""),
            ("مجموع فروخته شده:", total_sold),
            ("مجموع درآمد (ریال):", f"{total_revenue:,}"),
        ]
        
        for label, value in stats:
            ws.cell(row=row_num, column=1, value=label).font = ExcelExporter.SUMMARY_LABEL_FONT
            ws.cell(row=row_num, column=2, value=value).font = ExcelExporter.NORMAL_FONT
            row_num += 1
        
        row_num += 1  # فاصله
        
        # هدر جدول
        headers = [
            "شناسه", "نام محصول", "دسته‌بندی", "قیمت (ریال)", 
            "موجودی", "وضعیت فعال", "تعداد فروخته شده", "درآمد کل (ریال)"
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=header)
            cell.fill = ExcelExporter.HEADER_FILL
            cell.font = ExcelExporter.HEADER_FONT
            cell.alignment = ExcelExporter.HEADER_ALIGNMENT
        row_num += 1
        
        # داده‌های کامل
        for product in products:
            row = [
                product.get('id', ''),
                product.get('name', ''),
                product.get('category_name', ''),
                product.get('price', 0),
                product.get('stock_quantity', 0),
                'بله' if product.get('is_active') else 'خیر',
                product.get('total_sold') or 0,
                product.get('total_revenue') or 0
            ]
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.font = ExcelExporter.NORMAL_FONT
                cell.alignment = ExcelExporter.NORMAL_ALIGNMENT
            row_num += 1
        
        ExcelExporter.auto_adjust_column_width(ws)
        
        # ذخیره فایل در media و برگرداندن لینک
        if not filename:
            filename = f"product_report_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        file_url = ExcelExporter.save_excel_to_media(wb, filename, request)
        return file_url
    
    @staticmethod
    def export_stock_report(report_data: Dict[str, Any], filename: str = None, request=None) -> str:
        """Export stock report to Excel with complete data and statistics."""
        wb = ExcelExporter.create_workbook()
        ws = wb.active
        ws.title = "گزارش موجودی"
        
        # خلاصه آماری در ابتدا
        row_num = 1
        ws.merge_cells(f'A{row_num}:B{row_num}')
        summary_cell = ws.cell(row=row_num, column=1, value="خلاصه آماری گزارش موجودی")
        summary_cell.fill = ExcelExporter.SUMMARY_FILL
        summary_cell.font = ExcelExporter.SUMMARY_FONT
        summary_cell.alignment = Alignment(horizontal="center", vertical="center")
        row_num += 1
        
        stock_details = report_data.get('stock_details', [])
        low_stock_count = sum(1 for d in stock_details if d.get('is_low_stock'))
        out_of_stock_count = sum(1 for d in stock_details if d.get('is_out_of_stock'))
        active_count = sum(1 for d in stock_details if d.get('is_active'))
        
        # آمار کلی
        stats = [
            ("ارزش کل موجودی (ریال):", f"{report_data.get('total_stock_value', 0):,}"),
            ("تعداد کل آیتم‌ها:", report_data.get('total_items', 0)),
            ("تعداد کل محصولات:", len(stock_details)),
            ("محصولات فعال:", active_count),
            ("محصولات با موجودی کم:", low_stock_count),
            ("محصولات تمام شده:", out_of_stock_count),
        ]
        
        for label, value in stats:
            ws.cell(row=row_num, column=1, value=label).font = ExcelExporter.SUMMARY_LABEL_FONT
            ws.cell(row=row_num, column=2, value=value).font = ExcelExporter.NORMAL_FONT
            row_num += 1
        
        row_num += 1  # فاصله
        
        # هدر جدول
        headers = [
            "شناسه", "نام محصول", "دسته‌بندی", "موجودی", 
            "قیمت (ریال)", "ارزش موجودی (ریال)", "وضعیت فعال", "موجودی کم", "تمام شده"
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=header)
            cell.fill = ExcelExporter.HEADER_FILL
            cell.font = ExcelExporter.HEADER_FONT
            cell.alignment = ExcelExporter.HEADER_ALIGNMENT
        row_num += 1
        
        # داده‌های کامل
        for detail in stock_details:
            row = [
                detail.get('id', ''),
                detail.get('name', ''),
                detail.get('category_name', ''),
                detail.get('stock_quantity', 0),
                detail.get('price', 0),
                detail.get('stock_value', 0),
                'بله' if detail.get('is_active') else 'خیر',
                'بله' if detail.get('is_low_stock') else 'خیر',
                'بله' if detail.get('is_out_of_stock') else 'خیر'
            ]
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.font = ExcelExporter.NORMAL_FONT
                cell.alignment = ExcelExporter.NORMAL_ALIGNMENT
            row_num += 1
        
        ExcelExporter.auto_adjust_column_width(ws)
        
        # ذخیره فایل در media و برگرداندن لینک
        if not filename:
            filename = f"stock_report_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        file_url = ExcelExporter.save_excel_to_media(wb, filename, request)
        return file_url
    
    @staticmethod
    def export_daily_report(report_data: Dict[str, Any], filename: str = None, request=None) -> str:
        """Export daily report to Excel with complete data and statistics."""
        wb = ExcelExporter.create_workbook()
        ws = wb.active
        ws.title = "گزارش روزانه"
        
        # خلاصه آماری در ابتدا
        row_num = 1
        ws.merge_cells(f'A{row_num}:B{row_num}')
        summary_cell = ws.cell(row=row_num, column=1, value="خلاصه آماری گزارش روزانه")
        summary_cell.fill = ExcelExporter.SUMMARY_FILL
        summary_cell.font = ExcelExporter.SUMMARY_FONT
        summary_cell.alignment = Alignment(horizontal="center", vertical="center")
        row_num += 1
        
        total_orders = report_data.get('total_orders', 0)
        total_sales = report_data.get('total_sales', 0)
        total_transactions = report_data.get('total_transactions', 0)
        average_order = (total_sales / total_orders) if total_orders > 0 else 0
        
        # آمار کلی
        stats = [
            ("تاریخ گزارش:", report_data.get('date', '')),
            ("", ""),
            ("مجموع فروش (ریال):", f"{total_sales:,}"),
            ("تعداد کل سفارشات:", total_orders),
            ("میانگین ارزش هر سفارش (ریال):", f"{average_order:,.2f}"),
            ("تعداد تراکنش‌ها:", total_transactions),
        ]
        
        for label, value in stats:
            ws.cell(row=row_num, column=1, value=label).font = ExcelExporter.SUMMARY_LABEL_FONT
            ws.cell(row=row_num, column=2, value=value).font = ExcelExporter.NORMAL_FONT
            row_num += 1
        
        row_num += 1  # فاصله
        
        # هدر جدول (فیلدهای ضروری برای گزارش روزانه)
        headers = [
            "شماره سفارش", "مبلغ کل (ریال)", "وضعیت پرداخت", "تاریخ ایجاد"
        ]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_num, column=col_idx, value=header)
            cell.fill = ExcelExporter.HEADER_FILL
            cell.font = ExcelExporter.HEADER_FONT
            cell.alignment = ExcelExporter.HEADER_ALIGNMENT
        row_num += 1
        
        # داده‌های کامل (فقط فیلدهای ضروری)
        orders = report_data.get('orders', [])
        for order in orders:
            row = [
                order.get('order_number', ''),
                order.get('total_amount', 0),
                ExcelExporter.translate_payment_status(order.get('payment_status', '')),
                str(order.get('created_at', ''))[:19] if order.get('created_at') else ''
            ]
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.font = ExcelExporter.NORMAL_FONT
                cell.alignment = ExcelExporter.NORMAL_ALIGNMENT
            row_num += 1
        
        ExcelExporter.auto_adjust_column_width(ws)
        
        # ذخیره فایل در media و برگرداندن لینک
        if not filename:
            filename = f"daily_report_{report_data.get('date', timezone.now().strftime('%Y%m%d'))}.xlsx"
        
        file_url = ExcelExporter.save_excel_to_media(wb, filename, request)
        return file_url

